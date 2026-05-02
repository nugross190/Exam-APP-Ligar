"""End-to-end smoke test for Week 2 endpoints (spec §7 and §8).

Walks the teacher → admin → student path:

  Teacher portal (§7)
  1. List my exams
  2. List questions in one (empty)
  3. Create 3 questions: pg, tf, complex_mc (validates choice weights)
  4. Update one question (replaces choices wholesale)
  5. Upload an image to a question (writes to UPLOAD_DIR, returns /uploads/...)
  6. Negative: student token → 403; admin token → 404 (doesn't own subject)

  Admin panel (§8)
  7. POST /admin/exam/{id}/confirm — opens the exam
  8. Negative: after confirm, teacher mutations are frozen (400)
  9. Negative: student token on /admin/* → 403

  Student exam loop  (re-uses §5 / Week 2 day 1–2 work)
 10. Start session, fetch questions, answer each, submit
 11. Verify total/max scores match what we constructed

  Admin reporting / imports
 12. GET /admin/exam/{id}/monitor — counts include our submitted session
 13. GET /admin/results/{id}/export — returns an xlsx (PK-magic body)
 14. POST /admin/import/students?dry_run=true — parses without writing
 15. POST /admin/import/schedule?dry_run=true — parses without writing

The script is idempotent: at startup it scrubs the test exam's questions
and any prior session for the test student, resets admin_confirmed, and
ensures a known time window so /exam/start doesn't reject "exam window
has closed."

Run with:  python smoke_test_week2.py
Requires:  the DB has been seeded (see seed.py) and the test runs
           against the same DATABASE_URL as the FastAPI app.
"""
from __future__ import annotations

import io
import sys
from datetime import time, timedelta

import bcrypt
from fastapi.testclient import TestClient

from database import SessionLocal, utcnow
from main import app
from models import (
    AnswerChoice, Choice, Class, ClassSubject, Exam, ExamResult,
    ExamSession, ExpelledFlag, Question, SessionViolation, Student,
    StudentAnswer, Subject, Teacher, TeacherSubject,
)

client = TestClient(app)


# ---------------------------------------------------------------------------
# Helpers (same style as smoke_test.py)
# ---------------------------------------------------------------------------

def banner(s):
    print(f"\n{'='*70}\n  {s}\n{'='*70}")


def assert_eq(label, got, want):
    ok = got == want
    mark = "OK  " if ok else "FAIL"
    print(f"  [{mark}] {label}: got={got!r}  want={want!r}")
    if not ok:
        sys.exit(1)


def assert_true(label, cond, info=""):
    mark = "OK  " if cond else "FAIL"
    suffix = f"  ({info})" if info else ""
    print(f"  [{mark}] {label}{suffix}")
    if not cond:
        sys.exit(1)


# 1×1 transparent PNG, hand-crafted so the test doesn't depend on PIL.
_PNG_1x1 = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR"
    b"\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00\x00\x1f\x15\xc4"
    b"\x89\x00\x00\x00\rIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x01"
    b"\x00\xfb\xeeP\x88\x00\x00\x00\x00IEND\xaeB`\x82"
)


# ---------------------------------------------------------------------------
# Setup: pick an exam, a teacher (subject owner), an admin, a student.
# ---------------------------------------------------------------------------

banner("Setup: choose a test exam, teacher, admin, and student")
db = SessionLocal()
try:
    # Pick a subject that has at least one ClassSubject mapping AND a
    # non-flagged student in one of those classes. This guarantees the
    # student can /exam/start.
    candidates = (
        db.query(Subject)
        .join(ClassSubject, ClassSubject.subject_id == Subject.id)
        .join(Exam, Exam.subject_id == Subject.id)
        .all()
    )
    chosen_subject = None
    chosen_class = None
    chosen_student = None
    for subj in candidates:
        for cs in subj.class_subjects:
            stu = (
                db.query(Student)
                .filter_by(class_id=cs.class_id, flagged=False)
                .first()
            )
            if stu is not None:
                chosen_subject = subj
                chosen_class = cs.class_
                chosen_student = stu
                break
        if chosen_subject is not None:
            break
    assert_true(
        "found subject + class + non-flagged student",
        chosen_subject is not None,
    )
    print(f"  subject: {chosen_subject.name}")
    print(f"  class:   {chosen_class.name}")
    print(f"  student: {chosen_student.name}  nis={chosen_student.nis}")

    exam = db.query(Exam).filter_by(subject_id=chosen_subject.id).first()
    assert_true("exam exists for subject", exam is not None)

    # Get-or-create test teacher (subject owner). Fixed username so re-runs
    # don't multiply rows.
    teacher = db.query(Teacher).filter_by(username="smoke_teacher").first()
    if teacher is None:
        teacher = Teacher(
            username="smoke_teacher",
            password_hash=bcrypt.hashpw(b"teach123", bcrypt.gensalt(rounds=4)).decode(),
            full_name="Pak Guru Smoke",
            role="teacher",
        )
        db.add(teacher)
        db.flush()

    # Get-or-create test admin.
    admin = db.query(Teacher).filter_by(username="smoke_admin").first()
    if admin is None:
        admin = Teacher(
            username="smoke_admin",
            password_hash=bcrypt.hashpw(b"admin123", bcrypt.gensalt(rounds=4)).decode(),
            full_name="Bu Admin Smoke",
            role="admin",
        )
        db.add(admin)
        db.flush()

    # Wire teacher to the chosen subject so they own it. Authoring
    # ownership flows through TeacherSubject; the legacy
    # Subject.teacher_id is left as-is.
    existing_link = (
        db.query(TeacherSubject)
        .filter_by(teacher_id=teacher.id, subject_id=chosen_subject.id)
        .first()
    )
    if existing_link is None:
        db.add(TeacherSubject(
            teacher_id=teacher.id, subject_id=chosen_subject.id,
        ))

    # Reset exam state — admin_confirmed False so /admin/exam/{id}/confirm
    # has something to flip; window includes "now" so /exam/start works.
    now = utcnow()
    exam.admin_confirmed = False
    exam.status = "scheduled"
    exam.scheduled_at = now - timedelta(minutes=1)
    exam.time_end = time(23, 59)
    exam.duration_minutes = 90

    # Scrub prior data so the test is re-runnable. Cascades on Question
    # delete will clean up Choice + StudentAnswer + AnswerChoice.
    prior_q_ids = [q.id for q in exam.questions]
    if prior_q_ids:
        # Manually clear AnswerChoice + StudentAnswer first since they
        # don't cascade from Question (only from ExamSession).
        sa_ids = [
            sa.id for sa in db.query(StudentAnswer)
            .filter(StudentAnswer.question_id.in_(prior_q_ids)).all()
        ]
        if sa_ids:
            db.query(AnswerChoice).filter(
                AnswerChoice.student_answer_id.in_(sa_ids)
            ).delete(synchronize_session=False)
            db.query(StudentAnswer).filter(
                StudentAnswer.id.in_(sa_ids)
            ).delete(synchronize_session=False)
        db.query(Choice).filter(
            Choice.question_id.in_(prior_q_ids)
        ).delete(synchronize_session=False)
        db.query(Question).filter(
            Question.id.in_(prior_q_ids)
        ).delete(synchronize_session=False)

    prior_sessions = db.query(ExamSession).filter_by(exam_id=exam.id).all()
    prior_session_ids = [s.id for s in prior_sessions]
    if prior_session_ids:
        db.query(ExpelledFlag).filter(
            ExpelledFlag.session_id.in_(prior_session_ids)
        ).delete(synchronize_session=False)
        db.query(SessionViolation).filter(
            SessionViolation.session_id.in_(prior_session_ids)
        ).delete(synchronize_session=False)
        db.query(ExamResult).filter(
            ExamResult.session_id.in_(prior_session_ids)
        ).delete(synchronize_session=False)
        # AnswerChoice/StudentAnswer for these sessions
        sa_ids = [
            sa.id for sa in db.query(StudentAnswer)
            .filter(StudentAnswer.session_id.in_(prior_session_ids)).all()
        ]
        if sa_ids:
            db.query(AnswerChoice).filter(
                AnswerChoice.student_answer_id.in_(sa_ids)
            ).delete(synchronize_session=False)
            db.query(StudentAnswer).filter(
                StudentAnswer.id.in_(sa_ids)
            ).delete(synchronize_session=False)
        db.query(ExamSession).filter(
            ExamSession.id.in_(prior_session_ids)
        ).delete(synchronize_session=False)

    db.commit()

    test_exam_id = exam.id
    test_exam_title = exam.title
    test_student_id = chosen_student.id
    test_student_username = chosen_student.username
    test_student_password = chosen_student.nisn[-6:]
finally:
    db.close()


# ---------------------------------------------------------------------------
# Login all three actors.
# ---------------------------------------------------------------------------

banner("Login: teacher, admin, student")

r = client.post("/auth/teacher/login",
                json={"username": "smoke_teacher", "password": "teach123"})
assert_eq("teacher login status", r.status_code, 200)
T_HEADERS = {"Authorization": f"Bearer {r.json()['access_token']}"}

r = client.post("/auth/teacher/login",
                json={"username": "smoke_admin", "password": "admin123"})
assert_eq("admin login status", r.status_code, 200)
A_HEADERS = {"Authorization": f"Bearer {r.json()['access_token']}"}

r = client.post("/auth/student/login",
                json={"username": test_student_username,
                      "password": test_student_password})
assert_eq("student login status", r.status_code, 200)
S_HEADERS = {"Authorization": f"Bearer {r.json()['access_token']}"}


# ---------------------------------------------------------------------------
# §7.1  GET /teacher/exams
# ---------------------------------------------------------------------------

banner("§7.1  GET /teacher/exams")
r = client.get("/teacher/exams", headers=T_HEADERS)
assert_eq("status", r.status_code, 200)
exam_ids = [e["id"] for e in r.json()]
assert_true(f"test exam in teacher's list ({len(exam_ids)} total)",
            test_exam_id in exam_ids)


# Negative: admin doesn't own the subject, so list is empty.
r = client.get("/teacher/exams", headers=A_HEADERS)
assert_eq("admin's teacher/exams status", r.status_code, 200)
assert_true("admin sees no exams (no subjects assigned)",
            test_exam_id not in [e["id"] for e in r.json()])


# ---------------------------------------------------------------------------
# §7  GET /teacher/exam/{id}/questions  (empty after scrub)
# ---------------------------------------------------------------------------

banner("GET /teacher/exam/{id}/questions — initially empty")
r = client.get(f"/teacher/exam/{test_exam_id}/questions", headers=T_HEADERS)
assert_eq("status", r.status_code, 200)
assert_eq("question count", len(r.json()), 0)


# ---------------------------------------------------------------------------
# §7.2  POST /teacher/exam/{id}/question  — create three types
# ---------------------------------------------------------------------------

banner("§7.2  POST /teacher/exam/{id}/question — pg, tf, complex_mc")

# Q1 — pg (single-correct, 4 choices)
q1 = {
    "question_type": "pg",
    "body": "2 + 2 = ?",
    "item_points": 10,
    "choices_count": 4,
    "choices": [
        {"body": "3", "is_correct": False},
        {"body": "4", "is_correct": True},
        {"body": "5", "is_correct": False},
        {"body": "22", "is_correct": False},
    ],
}
r = client.post(f"/teacher/exam/{test_exam_id}/question", headers=T_HEADERS, json=q1)
assert_eq("pg create status", r.status_code, 201)
q1_id = r.json()["question_id"]

# Q2 — tf (single-correct, 2 choices)
q2 = {
    "question_type": "tf",
    "body": "Bumi datar.",
    "item_points": 5,
    "choices_count": 2,
    "choices": [
        {"body": "Benar", "is_correct": False},
        {"body": "Salah", "is_correct": True},
    ],
}
r = client.post(f"/teacher/exam/{test_exam_id}/question", headers=T_HEADERS, json=q2)
assert_eq("tf create status", r.status_code, 201)
q2_id = r.json()["question_id"]

# Q3 — complex_mc (multi-correct, 4 choices, 2 correct → weight 0.5 each)
q3 = {
    "question_type": "complex_mc",
    "body": "Pilih bilangan prima:",
    "item_points": 20,
    "choices_count": 4,
    "choices": [
        {"body": "2", "is_correct": True},
        {"body": "4", "is_correct": False},
        {"body": "5", "is_correct": True},
        {"body": "9", "is_correct": False},
    ],
}
r = client.post(f"/teacher/exam/{test_exam_id}/question", headers=T_HEADERS, json=q3)
assert_eq("complex_mc create status", r.status_code, 201)
q3_id = r.json()["question_id"]

# Verify weights — correct choices on q3 should have weight=0.5
r = client.get(f"/teacher/exam/{test_exam_id}/questions", headers=T_HEADERS)
qs = {q["id"]: q for q in r.json()}
q3_correct = [c for c in qs[q3_id]["choices"] if c["is_correct"]]
assert_eq("q3 correct choice count", len(q3_correct), 2)
assert_true("q3 weights == 0.5",
            all(abs(c["weight"] - 0.5) < 1e-9 for c in q3_correct))

# Invalid payload: complex_mc with 4 declared but 3 supplied
bad = {**q3, "choices": q3["choices"][:3]}
r = client.post(f"/teacher/exam/{test_exam_id}/question", headers=T_HEADERS, json=bad)
assert_eq("count-mismatch rejected", r.status_code, 400)

# pg with two correct choices → 400
bad2 = {**q1, "choices": [
    {"body": "3", "is_correct": True},
    {"body": "4", "is_correct": True},
    {"body": "5", "is_correct": False},
    {"body": "22", "is_correct": False},
]}
r = client.post(f"/teacher/exam/{test_exam_id}/question", headers=T_HEADERS, json=bad2)
assert_eq("pg multi-correct rejected", r.status_code, 400)


# ---------------------------------------------------------------------------
# §7.3  PUT /teacher/question/{id}
# ---------------------------------------------------------------------------

banner("§7.3  PUT /teacher/question/{id} — replace question + choices")
updated = {**q1, "body": "1 + 1 = ?", "choices": [
    {"body": "1", "is_correct": False},
    {"body": "2", "is_correct": True},
    {"body": "3", "is_correct": False},
    {"body": "11", "is_correct": False},
]}
r = client.put(f"/teacher/question/{q1_id}", headers=T_HEADERS, json=updated)
assert_eq("update status", r.status_code, 200)
assert_eq("body updated", r.json()["body"], "1 + 1 = ?")
assert_eq("choice count after update", len(r.json()["choices"]), 4)

# Negative: student token can't author
r = client.put(f"/teacher/question/{q1_id}", headers=S_HEADERS, json=updated)
assert_eq("student PUT rejected", r.status_code, 403)


# ---------------------------------------------------------------------------
# §7.4  POST /teacher/question/{id}/image
# ---------------------------------------------------------------------------

banner("§7.4  POST /teacher/question/{id}/image — upload PNG")
r = client.post(
    f"/teacher/question/{q1_id}/image",
    headers=T_HEADERS,
    files={"file": ("tiny.png", _PNG_1x1, "image/png")},
)
assert_eq("upload status", r.status_code, 200)
assert_true("image_url under /uploads/",
            r.json()["image_url"].startswith("/uploads/"))

# The static mount serves it back.
r = client.get(r.json()["image_url"])
assert_eq("static fetch status", r.status_code, 200)
assert_true("served bytes match", r.content == _PNG_1x1,
            f"len={len(r.content)} vs {len(_PNG_1x1)}")

# Negative: bad content-type
r = client.post(
    f"/teacher/question/{q1_id}/image",
    headers=T_HEADERS,
    files={"file": ("not.gif", b"GIF89a", "image/gif")},
)
assert_eq("gif rejected", r.status_code, 400)


# ---------------------------------------------------------------------------
# Negative: another teacher cannot author on someone else's exam
# ---------------------------------------------------------------------------

banner("Ownership: admin (no Subject.teacher_id) cannot author")
r = client.post(
    f"/teacher/exam/{test_exam_id}/question",
    headers=A_HEADERS, json=q1,
)
# admin's token passes get_current_teacher (role=admin is in the allowed
# set) but ownership check should 404 because Subject.teacher_id != admin.id.
assert_eq("non-owner gets 404", r.status_code, 404)


# ---------------------------------------------------------------------------
# §8.1  POST /admin/exam/{id}/confirm
# ---------------------------------------------------------------------------

banner("§8.1  POST /admin/exam/{id}/confirm")

# Negative: non-admin teacher token rejected
r = client.post(f"/admin/exam/{test_exam_id}/confirm", headers=T_HEADERS)
assert_eq("teacher rejected", r.status_code, 403)

# Negative: student token rejected
r = client.post(f"/admin/exam/{test_exam_id}/confirm", headers=S_HEADERS)
assert_eq("student rejected", r.status_code, 403)

# Happy path
r = client.post(f"/admin/exam/{test_exam_id}/confirm", headers=A_HEADERS)
assert_eq("admin confirm status", r.status_code, 200)
data = r.json()
assert_eq("admin_confirmed", data["admin_confirmed"], True)
# scheduled_at was set to now-1min, so status flips to 'open'
assert_eq("exam status", data["status"], "open")


# ---------------------------------------------------------------------------
# Frozen-after-confirm: teacher can no longer mutate
# ---------------------------------------------------------------------------

banner("Frozen-after-confirm: teacher mutations 400")
r = client.put(f"/teacher/question/{q1_id}", headers=T_HEADERS, json=updated)
assert_eq("PUT after confirm", r.status_code, 400)

r = client.post(f"/teacher/exam/{test_exam_id}/question", headers=T_HEADERS, json=q2)
assert_eq("POST after confirm", r.status_code, 400)


# ---------------------------------------------------------------------------
# Student exam loop — start, fetch, answer, submit
# ---------------------------------------------------------------------------

banner("Student exam loop (§5.1–§5.5)")

r = client.post("/exam/start", headers=S_HEADERS,
                json={"exam_id": test_exam_id})
assert_eq("start status", r.status_code, 200)
session_id = r.json()["session_id"]
assert_eq("questions_total", r.json()["questions_total"], 3)

r = client.get(f"/exam/{session_id}/questions", headers=S_HEADERS)
assert_eq("questions list status", r.status_code, 200)
qs_payload = r.json()["questions"]
assert_eq("got 3 questions", len(qs_payload), 3)
# Confirm choices are stripped of is_correct/weight
sample_choice = qs_payload[0]["choices"][0]
assert_true("is_correct hidden from student",
            "is_correct" not in sample_choice)
assert_true("weight hidden from student", "weight" not in sample_choice)

# Answer correctly. We need to look up which choice IDs are correct,
# but the student endpoint won't tell us — go to the DB instead, since
# this smoke test stands in for a teacher driving the UI.
db = SessionLocal()
try:
    correct_by_q: dict = {}
    for q in db.query(Question).filter_by(exam_id=test_exam_id).all():
        correct_by_q[q.id] = [c.id for c in q.choices if c.is_correct]
finally:
    db.close()

for q in qs_payload:
    answer_ids = correct_by_q[q["id"]]
    r = client.post(
        f"/exam/{session_id}/answer",
        headers=S_HEADERS,
        json={"question_id": q["id"], "choice_ids": answer_ids},
    )
    assert_eq(f"answer {q['question_type']:10s} status", r.status_code, 200)
    assert_eq(f"answer selected_count", r.json()["selected_count"],
              len(answer_ids))

r = client.post(f"/exam/{session_id}/submit", headers=S_HEADERS)
assert_eq("submit status", r.status_code, 200)
data = r.json()
expected_max = 10 + 5 + 20
assert_eq("max_score", data["max_score"], float(expected_max))
assert_eq("total_score (all correct)", data["total_score"], float(expected_max))
assert_eq("percentage", data["percentage"], 100.0)


# ---------------------------------------------------------------------------
# §8.2  GET /admin/exam/{id}/monitor
# ---------------------------------------------------------------------------

banner("§8.2  GET /admin/exam/{id}/monitor")
r = client.get(f"/admin/exam/{test_exam_id}/monitor", headers=A_HEADERS)
assert_eq("status", r.status_code, 200)
data = r.json()
assert_true("counts.total >= 1", data["counts"]["total"] >= 1,
            f"got {data['counts']}")
assert_true("counts.submitted >= 1", data["counts"]["submitted"] >= 1)
assert_eq("violations empty (we didn't violate)", data["violations"], [])
assert_eq("homeroom_flags empty", data["homeroom_flags"], [])


# ---------------------------------------------------------------------------
# §8.3  GET /admin/results/{id}/export
# ---------------------------------------------------------------------------

banner("§8.3  GET /admin/results/{id}/export — xlsx download")
r = client.get(f"/admin/results/{test_exam_id}/export", headers=A_HEADERS)
assert_eq("status", r.status_code, 200)
assert_true(
    "content-type is xlsx",
    "spreadsheetml" in r.headers.get("content-type", ""),
    info=r.headers.get("content-type", ""),
)
# .xlsx is a zip — magic bytes start with 'PK'
assert_true("body starts with PK (zip magic)", r.content[:2] == b"PK")

# Optional: verify openpyxl can read it back and find both sheets.
from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(r.content))
assert_true("has 'Per Student' sheet", "Per Student" in wb.sheetnames)
assert_true("has 'Per Class' sheet", "Per Class" in wb.sheetnames)
ws = wb["Per Student"]
assert_true(f"Per Student has >= 2 rows (header + data; got {ws.max_row})",
            ws.max_row >= 2)


# ---------------------------------------------------------------------------
# §8.4  POST /admin/import/students  (dry_run)
# ---------------------------------------------------------------------------

banner("§8.4  POST /admin/import/students?dry_run=true")
with open("daftar_peserta_kelas_XI_updated.xlsx", "rb") as fxi, \
     open("daftar_peserta_kelas_X_updated.xlsx", "rb") as fx:
    r = client.post(
        "/admin/import/students?dry_run=true",
        headers=A_HEADERS,
        files={
            "xi_file": ("xi.xlsx", fxi.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            "x_file":  ("x.xlsx",  fx.read(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
    )
assert_eq("status", r.status_code, 200)
data = r.json()
assert_eq("dry_run", data["dry_run"], True)
assert_true("parsed_rows in expected ballpark (>=800)",
            data["parsed_rows"] >= 800, info=str(data["parsed_rows"]))
assert_true("seed_stats absent in dry_run", data.get("seed_stats") is None)


# ---------------------------------------------------------------------------
# §8.5  POST /admin/import/schedule  (dry_run)
# ---------------------------------------------------------------------------

banner("§8.5  POST /admin/import/schedule?dry_run=true")
with open("schedule_parsed.csv", "rb") as fs:
    r = client.post(
        "/admin/import/schedule?dry_run=true",
        headers=A_HEADERS,
        files={"schedule_file": ("schedule.csv", fs.read(), "text/csv")},
    )
assert_eq("status", r.status_code, 200)
data = r.json()
assert_eq("dry_run", data["dry_run"], True)
assert_true("parsed_entries > 0", data["parsed_entries"] > 0,
            info=str(data["parsed_entries"]))


# ---------------------------------------------------------------------------
# Negative: student token can't reach /admin/*
# ---------------------------------------------------------------------------

banner("Authz: student token blocked from /admin/*")
r = client.get(f"/admin/exam/{test_exam_id}/monitor", headers=S_HEADERS)
assert_eq("monitor rejected", r.status_code, 403)
r = client.get(f"/admin/results/{test_exam_id}/export", headers=S_HEADERS)
assert_eq("export rejected", r.status_code, 403)


banner("ALL WEEK 2 SMOKE TESTS PASSED")
